import openpyxl
import xlrd
from commons.yaml_util import get_object_path, write_extract_yaml
import os

#读取excel中的指定sheet页的全部数据
def read_excel(Sheet1, path):
    """
    读取excel中的指定sheet页的全部数据,此第三方库不支持xls后缀文件。
    读取excel中的指定sheet页的全部数据
    :param Sheet1: excel文件中的sheet页名称
    :param path: excel文件的路径
    :return: 返回excel文件中指定的sheet页数据
    """
    # 通过路径获取工作簿
    wk = openpyxl.load_workbook(path)
    # 获取工作簿中的sheet页
    my_sheet = wk[Sheet1]

    datas = []
    # 从excel文件Sheet1工作表一次性读取所有的数据？
    # 循环my_sheet工作表中每一行
    for row in range(2, my_sheet.max_row + 1):
        rowdata = {}
        # 3...n列，3个单元格数据
        # 循环每行中的每一列
        for col in range(1, my_sheet.max_column + 1):
            key = my_sheet.cell(1, col).value
            value = my_sheet.cell(row, col).value
            rowdata[key] = value
        datas.append(rowdata)
    return datas


#获取excel文件中指定sheet页的指定行数据
def get_sheet_row(num,Sheet1, path):
    """
    获取excel文件中指定sheet页的指定行数据
    :param num: 指定行数据
    :param Sheet1: excel文件中的sheet页名称
    :param path: excel文件的路径
    :return: 返回excel文件中指定sheet页的指定行数据
    """
    row = read_excel(Sheet1, path)[num-1]
    return row


#通过sheet页名称读取excel中的指定sheet页的全部数据
def get_Excel_All(filename,sheetname):
    """
    过sheet页名称读取excel中的指定sheet页的全部数据
    :param filename: 文件路径及名称
    :param sheetname: sheet名称
    :return: 返回对应sheet页列的表数据
    """
    xl_name = xlrd.open_workbook(filename)
    xl_sheet = xl_name.sheet_by_name(sheetname)
    datalist = []
    for x in range(0,xl_sheet.nrows):
        lst = []
        for y in range(xl_sheet.ncols):
            value = xl_sheet.row_values(x)[y]
            if type(value) == float:
                value = int(value)
            lst.append(value)
        datalist.append(tuple(lst))
    return datalist



#通过sheet索引读取excel中的索引的全部数据
def get_Excel_All_index(filename,index=0):
    """
    通过sheet索引读取excel中的索引的全部数据，默认索引为第一个
    :param filename: 文件路径及名称
    :param index: sheet页索引，默认为0，第一个sheet页
    :return: 返回对应sheet页的列表数据
    """

    xl_name = xlrd.open_workbook(filename)
    xl_sheet = xl_name.sheet_by_index(index)
    datalist = []
    for x in range(0,xl_sheet.nrows):
        lst = []
        for y in range(xl_sheet.ncols):
            value = xl_sheet.row_values(x)[y]
            if type(value) == float:
                value = int(value)
            lst.append(value)
        datalist.append(tuple(lst))
    return datalist

# 获取目录下所有文件夹中的所有excel表
def get_all_tables(dir_name):
    """
    获取目录下所有文件夹中的所有excel表
    :param dir_name: 指定文件夹目录
    :return: 返回指定文件夹目录下的所有excel表格-列表
    """
    tables = []
    for excel in os.listdir(get_object_path() + '/'+ f"{dir_name}"):  # os.listdir() 方法用于返回指定的文件夹包含的文件或文件夹的名字的列表
        if excel.endswith('xlsx') or excel.endswith('xls'):
            excel_path = os.path.join(get_object_path() + '/'+ f"{dir_name}",excel) # 拼接路径
            # excels = xlrd.open_workbook(excel_path) #打开路径的表格
            tables.append(excel_path)
            # sheet_names = excels.sheet_names()   #获取目录下所有Sheet的名字，为list
    return tables

#获取目录下的所有表格中的所有sheet页名称
def get_all_sheet(dir_name):
    """
    获取目录下的所有表格中的所有sheet页名称
    :param dir_name: 指定文件夹目录
    :return: 返回目录下的所有表格中的所有sheet页名称
    """
    lis = []
    for table in get_all_tables(dir_name):
        excels = xlrd.open_workbook(table)
        sheet_names = excels.sheet_names()
        for sheet in sheet_names:
            lis.append((sheet,table))
            # print(read_excel(sheet, table))
    return lis


#读取目录下的所有表格中的所有sheet页数据
def get_all_caseinfo(dir_name="caseInfo"):
    """
    读取目录下的所有表格中的所有sheet页数据
    :return: 返回目录下的所有表格中的所有sheet页数据
    """
    #调用上面方法获取所有sheet页名称
    datas = get_all_sheet(dir_name)

    lis = []
    for data in datas:
        caseinfos = read_excel(data[0], data[1])  #只支持xlsx
        # caseinfos = get_Excel_All(data[1], data[0])

        for data in caseinfos:
            lis.append(data)
    return lis

#读取excel中的指定sheet页的全部数据(数据驱动)
def read_data_excel(path):
    """
    读取excel中的指定sheet页的全部数据,此第三方库不支持xls后缀文件。
    读取excel中的指定sheet页的每行的数据
    :param Sheet1: excel文件中的sheet页名称
    :param path: excel文件的路径
    :return: 返回excel每行列表数据
    """
    # 通过路径获取工作簿
    wb = xlrd.open_workbook(path)
    # 获取工作簿中的sheet页
    table = wb.sheets()[0]
    nrows = table.nrows  # 行数
    ncols = table.ncols  # 列数
    datas = []
    for i in range(0, nrows):
        row_list = []
        rowValues = table.row_values(i)  # 某一行数据
        for data in rowValues:
            if str(data).endswith(".0"):
                data = int(data)
            row_list.append(data)
        datas.append(row_list)
    return datas

    # datas = []
    # # 从excel文件Sheet1工作表一次性读取所有的数据？
    # # 循环my_sheet工作表中每一行
    # for row in range(2, nrows ):
    #     rowdata = {}
    #     # 3...n列，3个单元格数据
    #     # 循环每行中的每一列
    #     for col in range(1, ncols + 1):
    #         key = table.cell(1, col). value
    #         value = table.cell(row, col).value
    #         rowdata[key] = value
    #     datas.append(rowdata)



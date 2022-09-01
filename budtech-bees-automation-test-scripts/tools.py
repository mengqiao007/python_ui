import json
from openpyxl import Workbook
import commons.mysql_util as mysql
import requests
import jsonpath
import commons.function_util as function_util
import threading


# 中台用户类
class MP_TOOLS():
    """
    environment=0 ：dev环境； 1 ：test环境； 2 ：feature-2环境； 不填默认为0： DEV环境
    """

    def __init__(self, environment=0):  # 0=dev环境， 1=test环境, 2=feature-2环境
        self.environment = environment

    # 执行创建工号操作
    def mp_mysql_get_job_number(self, job_number, phone_number, name, parent_job_number, email, servie="MYSQL_MP"):
        """
        通过sql语句执行创建账号操作
        :param job_number: 创建用户新的工号
        :param phone_number: 用户手机号
        :param name: 用户名
        :param parent_job_number: 用户上级工号
        :param email: 用户邮箱
        """

        execute_mysql = f"INSERT INTO `mp_people_user` (`user_type`, `employee_no`, " \
                        f"`phone`, `real_name`, `real_en_name`, `real_name_visibility`, `manager_emp_no`, `manager_name`, " \
                        f"`on_board_date`, `email`, `status`, `gender`, `birth_date`, `sap_position`, `position`, `area_type`, " \
                        f"`channel_code`, `ignite_id`, `create_by`, `update_by`, `deleted`) " \
                        f"VALUES ( 1, '{job_number}', '{phone_number}', '{name}', '', 1,'{parent_job_number}', 'TEST', None, " \
                        f"'{email}', 3, 3, None, '', 'M1', 1, 'TEST', '', 'MP-TEST', 'MP-TEST',0);"
        dataname = ""
        if self.environment == 0:
            dataname = "abi-cloud-middle-platform-user-dev"
        elif self.environment == 1:
            dataname = "abi-cloud-middle-platform-user-test"
        elif self.environment == 2:
            dataname = "abi-cloud-middle-platform-user-dev"

        # 执行sql语句
        try:
            mysql.mysql_update(servie, dataname, execute_mysql)
        except Exception as e:
            print("sql语句创建工号错误")
            raise e

    # 创建工号-创建中台账号-提取密码
    def mp_create_account(self, job_number, phone_number, name, parent_job_number, email, servie="MYSQL_MP"):
        """
        先查询工号是否存在，创建对应环境工号，通过工号创建中台账号，提取中台账号对应密码，返回中台userid
        :param job_number: 创建用户新的工号
        :param phone_number: 用户手机号
        :param name: 用户名
        :param parent_job_number: 用户上级工号
        :param email: 用户邮箱
        :param return: 返回创建的工号，密码及中台用户id
        """
        # 查询工号是否存在
        if self.select_job_number_exist(job_number) != None:

            return "此工号已存在，不能再次创建"
        else:
            # 调用方法生成工号
            self.mp_mysql_get_job_number(job_number, phone_number, name, parent_job_number, email)

            # 通过api使用工号创建中台账号，并返回userid
            userid = self.mp_account_api(job_number)

            # 调用方法获取密码
            password = self.mp_mysql_get_password(service=servie, email=email)

            dic = {}
            dic["工号"] = job_number
            dic["密码"] = password
            dic["中台用户id"] = userid
            return dic

    # 查询邮箱对应的密码
    def mp_mysql_get_password(self, email, service):
        """
        :param service: 选择数据连接的服务
        :param email: 查询的邮箱
        :return: 通过服务和邮箱参数，执行sql语句查出对应的密码
        """

        execute_mysql = f'SELECT * from communication_email_send_log WHERE receiver_email_address="{email}" ORDER BY create_time DESC'
        dataname = ""
        if self.environment == 0 or self.environment == 2:
            dataname = "abi_cloud_middle_platform_communication_dev"
        elif self.environment == 1:
            dataname = "abi_cloud_middle_platform_communication_test"
        res = mysql.mysql_select_one(service, dataname, execute_mysql)
        password = res["message_content"][12:]
        return password

    # 根据工号创建中台账号
    def mp_account_api(self, job_number):
        """
        根据工号创建中台账号
        :param job_number: 已存在的工号
        :return: 返回对应的中台userid
        """
        # function_util.FunctionUtil().get_mp_login_token()

        base_url = ""
        params = ""
        header = ""
        if self.environment == 0:
            base_url = "https://api-gateway-dev.ab-inbev.cn"
            params = {"employeeNo": f"{job_number}", "roleCodes": ["10081", "10131", "00100001"],
                      "moduleCodes": ["1", "2", "3", "4", "6"]}
            header = {"satoken": function_util.FunctionUtil().get_mp_login_token("MP_DEV_CONSOLE_TOKEN")}
        elif self.environment == 1:
            base_url = "https://api-gateway-test.ab-inbev.cn"
            header = {"satoken": function_util.FunctionUtil().get_mp_login_token("MP_TEST_CONSOLE_TOKEN")}
            params = {"employeeNo": f"{job_number}", "roleCodes": ["11031", "11073", "10312", "00100001", "10862"],
                      "moduleCodes": ["1", "2", "3", "4", "6"]}
        elif self.environment == 2:
            base_url = "https://api-gateway-feature-2.ab-inbev.cn"
            header = {"satoken": function_util.FunctionUtil().get_mp_login_token("MP_FEATURE2_CONSOLE_TOKEN")}
            params = {"employeeNo": f"{job_number}", "roleCodes": ["10081", "10131", "00100001", "10021"],
                      "moduleCodes": ["1", "2", "3", "4", "6"]}

        url = base_url + f"/abi-cloud-middle-platform-user-service/console/system/user/add"
        res = requests.post(url=url, headers=header, json=params)
        userid = jsonpath.jsonpath(res.json(), "$..userLoginId")
        return userid[0]

    # 查询工号是否存在
    def select_job_number_exist(self, job_number, servie="MYSQL_MP"):
        """
        查询工号是否存在
        :param job_number: 查询的工号
        :param servie: 选择服务
        :return: 返回查询到的结果，不存在为None，存在及返回数据
        """
        execute_mysql = f"select * from mp_user_employee where employee_no='{job_number}'"
        dataname = ""
        if self.environment == 0:
            dataname = "abi-cloud-middle-platform-user-dev"
        elif self.environment == 1:
            dataname = "abi-cloud-middle-platform-user-test"
        elif self.environment == 2:
            dataname = "abi-cloud-middle-platform-user-dev"

        res = mysql.mysql_select_one(servie, dataname, execute_mysql)
        return res

    # 对postman脚本的数据进行格式化生成到excel文件中
    def postman_generate_excel(self, json_files_path, save_files_path):
        """
        对postman脚本的数据进行格式化生成到excel文件中
        :param json_files_path: postman脚本的json文件路径
        :param save_files_path: 需要生成excel的文件路径
        :return:
        """
        # 读取json文件数据
        with open(json_files_path, "r", encoding="utf-8") as f:
            datas = json.load(f)

        # 加载excel表格
        wb = Workbook()
        server_item = datas["item"]

        # 激活excel表
        for i in range(len(server_item)):
            module = server_item[i]["name"]

            self.create_sheet(wb, module, i, datas)
        wb.save(save_files_path)

    # 创建sheet页
    def create_sheet(self, wb, sheet_name, sheet_index, datas):
        """
        创建sheet页
        :param wb: 工作簿对象
        :param sheet_name: sheet名称
        :param sheet_index: sheet索引/遍历索引值
        :param datas: 传入的json文件
        :return: 写入excel
        """
        # 激活excel表

        sheet = wb.create_sheet(sheet_name, sheet_index)
        sheet["a1"] = "server"
        sheet["b1"] = "module"
        sheet["c1"] = "title"
        sheet["d1"] = "method"
        sheet["e1"] = "mode"
        sheet["f1"] = "body"
        sheet["g1"] = "url"

        # 服务名称
        server = datas["info"]["name"]
        sheet["a2"] = server

        # 获取服务数据
        server_item = datas["item"]

        # 模块名
        module = server_item[sheet_index]["name"]
        module_item = server_item[sheet_index]["item"]

        sheet[f"b2"] = module

        # 遍历模块下的所有接口用例
        for i in range(len(module_item)):
            title = module_item[i]["name"]
            request = module_item[i]["request"]
            method = request["method"]
            mode = request["body"]["mode"]
            # 判断body类型
            if "raw" in request["body"]:
                body = request["body"]["raw"]
            elif "urlencoded" in request["body"]:
                if request["body"]["urlencoded"] != []:
                    dic = {}
                    for ib in request["body"]["urlencoded"]:
                        a = ib["key"]
                        b = ib["value"]
                        dic[a] = b
                    body = str(dic)
                else:
                    body = "参数为空"
            else:
                body = "body不是raw格式，请手动填写"
            url = request["url"]["raw"]

            # 写入数据
            sheet[f"c{2 + i}"] = title
            sheet[f"d{2 + i}"] = method
            sheet[f"e{2 + i}"] = mode
            sheet[f"f{2 + i}"] = body
            sheet[f"g{2 + i}"] = url

    def send_email_reset_password(self):

        base_url = ""
        if self.environment == 0:
            base_url = "https://api-gateway-dev.ab-inbev.cn"
            header = {"satoken": function_util.FunctionUtil().get_mp_login_token("MP_DEV_CONSOLE_TOKEN")}
        elif self.environment == 1:
            base_url = "https://api-gateway-test.ab-inbev.cn"
            header = {"satoken": function_util.FunctionUtil().get_mp_login_token("MP_TEST_CONSOLE_TOKEN")}
        elif self.environment == 2:
            base_url = "https://api-gateway-feature-2.ab-inbev.cn"
            header = {"satoken": function_util.FunctionUtil().get_mp_login_token("MP_FEATURE2_CONSOLE_TOKEN")}

        url = base_url + "/abi-cloud-middle-platform-user-service/console/password/reset/send-verification-code/email"
        params = {"receiverEmailAddress":"mengqiao.wu@budweiserapac.com"}
        res = requests.get(url=url,params=params)
        return res.json()

if __name__ == '__main__':
    # print(MP_TOOLS(1).mp_create_account(job_number="28000110", phone_number=15801775931, name="于倩",
    #                                    parent_job_number="28073807",
    #                                    email="qian.yu@budweiserapac.com"))

    # MP_TOOLS().postman_generate_excel("./files/DAG.json","./files/DAG_1.xlsx")

    # print(function_util.FunctionUtil().get_localtime_str())
    print(MP_TOOLS().send_email_reset_password())



import random
import openpyxl
from commons.yaml_util import write_extract_yaml, read_environment_yaml
import requests
import jsonpath
from commons.logging_util import loggingUtil
from commons.function_util import FunctionUtil


# 前置方法类
class PrepositionUtil():

    def product_preposition(self):
        """
        商品服务前置数据
        @return:
        """
        num = FunctionUtil().random_code(3)
        #品牌
        self.product_brand_data(num,"Brand")
        self.product_brand_data(num,"Category")
        self.product_brand_data(num, "Unit")


    def product_brand_data(self,num,name):
        str_name = name.lower()
        if str_name == "unit":
            self.write_extract_yaml_format(f"{str_name}_name", f"add{name}" + num)
            self.write_extract_yaml_format(f"{str_name}_edit_name", f"edit{name}" + num)
        else:
            self.write_extract_yaml_format(f"{str_name}_name", f"add{name}" + num)
            self.write_extract_yaml_format(f"{str_name}_edit_name", f"edit{name}" + num)
            self.write_extract_yaml_format(f"child_{str_name}_name", f"addChild{name}" + num)
            self.write_extract_yaml_format(f"child_{str_name}_edit_name", f"editChild{name}" + num)



    def store_batch_delete_data(self):
        """
        店铺模块批量删除前置新增店铺数据
        @return:
        """
        url = read_environment_yaml("business_entity_service") + "store"
        header_abi = {"authentoken": read_environment_yaml("bees_console_token_abi"),
                      "type" : "1"}
        num = str(random.randint(1, 99999)).zfill(5)
        name = "auto-test-" + num
        params = {
                      "storeCode": "1233212",
                      "storeDetails": "1233212",
                      "storeName": name,
                      "wholesalerId": 1334065882783074539
                 }
        res = requests.post(url,json=params,headers=header_abi)
        status = jsonpath.jsonpath(res.json(),"$..status")[0]
        if status != "SUCCESS":
            loggingUtil().get_error_log("店铺模块前置，创建店铺错误,返回结果："+ str(res.json()))

        select_url = read_environment_yaml("business_entity_service") + "store/paging"
        select_params = {"storeName":name}
        select_res = requests.get(select_url,params=select_params,headers=header_abi)
        try:
            store_id = jsonpath.jsonpath(select_res.json(),"$..id")[0]
        except Exception as e:
            loggingUtil().get_error_log("店铺模块前置，查找店铺id异常,返回结果："+str(select_res.json()))
            raise e
        return store_id,name


    def store_batch_delete_id(self):

        for i in range(2):
            store_id, name = self.store_batch_delete_data()
            self.write_extract_yaml_format(f"store_id_{i+1}",store_id)
            self.write_extract_yaml_format(f"store_name_{i+1}",name)



    def write_extract_yaml_format(self,key,value):
        """
        写入yaml文件方法
        @param key: key
        @param value: value
        @return: 写入yaml文件
        """
        dic = {key:value}
        write_extract_yaml(dic)


    def poc_group_format_data(self):
        """
        poc_group前置数据
        @return:
        """
        wholesaler_id = {"group_wholesaler_id":"1334065882783073822"}
        add_poc_id = {"group_add_poc_id":"1334065882783135865"}
        write_extract_yaml(wholesaler_id)
        write_extract_yaml(add_poc_id)


    def wholesaler_format_data(self):
        """
        经销商导入前置数据
        @return:
        """
        self.wholesaler_format_data_()
        self.abi_import_t1_wholesaler_excel_data_format()
        self.abi_import_t2_wholesaler_excel_data_format()
        self.t1_import_t2_wholesaler_excel_data_format()


    def wholesaler_format_data_(self):
        num = str(random.randint(1, 99999)).zfill(5)

        wholesaler_format_name = "AddWholesalerX" + num
        wholesaler_format_phone = "133777" + num

        wholesaler_format_name = {"wholesaler_format_name": wholesaler_format_name}
        wholesaler_format_phone = {"wholesaler_format_phone": wholesaler_format_phone}
        write_extract_yaml(wholesaler_format_name)
        write_extract_yaml(wholesaler_format_phone)

    def t1_import_t2_wholesaler_excel_data_format(self,path="./files/t1_import_t2_wholesaler_data_format.xlsx",
                                                  sheet="sheet1"):
        # 通过路径获取工作簿
        wk = openpyxl.load_workbook(path)
        # 获取工作簿中的sheet页
        my_sheet = wk[sheet]

        for i in range(1):
            num = str(random.randint(1, 99999)).zfill(5)
            province = "山东省"
            city = "青岛市"
            county = "市南区"
            address = "金门路"
            name_connect = "T1importT2x" + num
            phone_connect = int("133106" + num)
            job_num = "78100001"
            t1_job_num = "w0000764"
            grant_web = "是"
            my_sheet[f"A{i + 2}"] = name_connect
            my_sheet[f"B{i + 2}"] = name_connect
            my_sheet[f"C{i + 2}"] = phone_connect
            my_sheet[f"D{i + 2}"] = province
            my_sheet[f"E{i + 2}"] = city
            my_sheet[f"F{i + 2}"] = county
            my_sheet[f"G{i + 2}"] = address
            my_sheet[f"H{i + 2}"] = province
            my_sheet[f"I{i + 2}"] = city
            my_sheet[f"J{i + 2}"] = county
            my_sheet[f"K{i + 2}"] = address
            my_sheet[f"L{i + 2}"] = name_connect
            my_sheet[f"M{i + 2}"] = name_connect
            my_sheet[f"N{i + 2}"] = name_connect
            my_sheet[f"O{i + 2}"] = phone_connect
            my_sheet[f"R{i + 2}"] = job_num
            my_sheet[f"S{i + 2}"] = t1_job_num
            my_sheet[f"AC{i + 2}"] = grant_web
            select_wholesaler = {"T1importT2_name": name_connect}
            write_extract_yaml(select_wholesaler)
        wk.save(path)

    def abi_import_t1_wholesaler_excel_data_format(self,path="./files/abi_import_t1_wholesaler_data_format.xlsx",
                                                   sheet="Sheet1"):
        # 通过路径获取工作簿
        wk = openpyxl.load_workbook(path)
        # 获取工作簿中的sheet页
        my_sheet = wk[sheet]

        for i in range(1):
            num = str(random.randint(1, 99999)).zfill(5)
            name_connect = "ABIimportT1x" + num
            phone_connect = int("133107" + num)
            my_sheet[f"A{i + 2}"] = name_connect
            my_sheet[f"B{i + 2}"] = name_connect
            my_sheet[f"C{i + 2}"] = "一批"
            my_sheet[f"D{i + 2}"] = name_connect
            my_sheet[f"E{i + 2}"] = name_connect
            my_sheet[f"F{i + 2}"] = phone_connect
            my_sheet[f"G{i + 2}"] = name_connect
            my_sheet[f"H{i + 2}"] = name_connect
            my_sheet[f"I{i + 2}"] = phone_connect
            select_wholesaler = {"delete_wholesaler_name": name_connect}
            write_extract_yaml(select_wholesaler)
        wk.save(path)

    def abi_import_t2_wholesaler_excel_data_format(self,path="./files/abi_import_t2_wholesaler_data_format.xlsx",
                                                   sheet="sheet1"):
        # 通过路径获取工作簿
        wk = openpyxl.load_workbook(path)
        # 获取工作簿中的sheet页
        my_sheet = wk[sheet]

        for i in range(1):
            num = str(random.randint(1, 99999)).zfill(5)
            t1_wholersaler = "Subtributor"
            province = "山东省"
            city = "青岛市"
            county = "市南区"
            address = "金门路"
            name_connect = "ABIimportT2x" + num
            phone_connect = int("133106" + num)
            job_num = "78100001"
            t1_job_num = "w0000764"
            my_sheet[f"A{i + 2}"] = t1_wholersaler
            my_sheet[f"B{i + 2}"] = name_connect
            my_sheet[f"C{i + 2}"] = name_connect
            my_sheet[f"D{i + 2}"] = phone_connect
            my_sheet[f"E{i + 2}"] = province
            my_sheet[f"F{i + 2}"] = city
            my_sheet[f"G{i + 2}"] = county
            my_sheet[f"H{i + 2}"] = address
            my_sheet[f"I{i + 2}"] = province
            my_sheet[f"J{i + 2}"] = city
            my_sheet[f"K{i + 2}"] = county
            my_sheet[f"L{i + 2}"] = address
            my_sheet[f"M{i + 2}"] = name_connect
            my_sheet[f"N{i + 2}"] = name_connect
            my_sheet[f"O{i + 2}"] = name_connect
            my_sheet[f"P{i + 2}"] = phone_connect
            my_sheet[f"S{i + 2}"] = job_num
            my_sheet[f"T{i + 2}"] = t1_job_num
            select_wholesaler1 = {"ABIimportT2_name": name_connect}
            select_wholesaler = {"delete_warehouse_name": name_connect}
            write_extract_yaml(select_wholesaler1)
            write_extract_yaml(select_wholesaler)
        wk.save(path)

